from openpyxl import Workbook
import openpyxl as xl
from openpyxl.chart import (
    AreaChart,
    Reference,
    Series,
)
from openpyxl.chart.reference import Reference
from openpyxl.chart.bar_chart import BarChart

wb = Workbook()
ws = wb.active

rows = [
    ['Number', 'Batch 1', 'Batch 2', 'Points'],
    [2, 40, 30, 9],
    [3, 40, 25, 8],
    [4, 50, 30, 8],
    [5, 30, 10, 7],
    [6, 25, 5, 10],
    [7, 50, 10, 8],
]
# create data in excell
for row in rows:
    ws.append(row)
'''
wb1 = xl.load_workbook('/home/tuandinh/Desktop/python/area.xlsx')
sheet = wb1['Sheet']

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 4)
    corrected_point = cell.value * 0.9
    corrected_point_cell = sheet.cell(row, 5)
    corrected_point_cell.value = corrected_point
'''
chart = BarChart()
chart.title = "Bar Chart"
chart.style = 25
data = Reference(ws,
                min_row=2,
                max_row= ws.max_row,
                min_col=5,
                max_col=5)

chart.add_data(data)
ws.add_chart(chart, 'A10')


chart = AreaChart()
chart.title = "Area Chart"
chart.style = 13
chart.x_axis.title = 'Test'
chart.y_axis.title = 'Percentage'

cats = Reference(ws, min_col=1, min_row=1, max_row=7)
data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws.add_chart(chart, "M10")

wb.save("area.xlsx")